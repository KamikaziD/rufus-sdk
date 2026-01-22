import httpx
from fastapi import UploadFile
from typing import List, Dict, Any, Optional
import magic
import pypdf
from docx import Document
import openpyxl
import io
import os
import platform
import base64

if platform.system() == "Darwin":
    # For macOS, try to find libmagic installed by Homebrew
    brew_libmagic_path = "/usr/local/Cellar/libmagic/5.46/lib/libmagic.dylib"
    if os.path.exists(brew_libmagic_path):
        os.environ["MAGIC_LIB"] = brew_libmagic_path

class FileService:
    async def get_files_from_urls(self, urls: List[str]) -> List[Dict[str, Any]]:
        """Fetch files from a list of URLs."""
        files = []
        async with httpx.AsyncClient(timeout=300.0) as client:
            for url in urls:
                try:
                    response = await client.get(url)
                    response.raise_for_status()
                    files.append({
                        "filename": url.split("/")[-1],
                        "content": response.content
                    })
                except httpx.RequestError as e:
                    print(f"Error fetching {url}: {e}")
        return files

    async def get_files_from_uploads(self, uploaded_files: List[UploadFile]) -> List[Dict[str, Any]]:
        """Get files from a list of uploaded files."""
        files = []
        for file in uploaded_files:
            content = await file.read()
            files.append({
                "filename": file.filename,
                "content": content
            })
        return files

    def read_file_content(self, file: Dict[str, Any]) -> Dict[str, str]:
        """Read the content of a file and return it as a dictionary with type and content."""
        mime_type = magic.from_buffer(file["content"], mime=True)
        
        if "pdf" in mime_type:
            return {"type": "text", "content": self._read_pdf(file["content"])}
        elif "vnd.openxmlformats-officedocument.wordprocessingml.document" in mime_type:
            return {"type": "text", "content": self._read_docx(file["content"])}
        elif "vnd.openxmlformats-officedocument.spreadsheetml.sheet" in mime_type:
            return {"type": "text", "content": self._read_xlsx(file["content"])}
        elif "text" in mime_type:
            return {"type": "text", "content": file["content"].decode("utf-8")}
        elif mime_type in ["image/jpeg", "image/png"]:
            return {"type": "image", "content": base64.b64encode(file["content"]).decode("utf-8")}
        else:
            return {"type": "unsupported", "content": f"Unsupported file type: {mime_type}"}

    def _read_pdf(self, content: bytes) -> str:
        """Read the content of a PDF file."""
        try:
            pdf_reader = pypdf.PdfReader(io.BytesIO(content))
            text = ""
            for page in pdf_reader.pages:
                text += page.extract_text()
            return text
        except Exception as e:
            return f"Error reading PDF: {e}"

    def _read_docx(self, content: bytes) -> str:
        """Read the content of a DOCX file."""
        try:
            doc = Document(io.BytesIO(content))
            return "\n".join([para.text for para in doc.paragraphs])
        except Exception as e:
            return f"Error reading DOCX: {e}"

    def _read_xlsx(self, content: bytes) -> str:
        """Read the content of an XLSX file."""
        try:
            workbook = openpyxl.load_workbook(io.BytesIO(content))
            text = ""
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                for row in sheet.iter_rows():
                    text += "\t".join([str(cell.value) for cell in row]) + "\n"
            return text
        except Exception as e:
            return f"Error reading XLSX: {e}"

file_service = FileService()

